import sqlite3, os, secrets, smtplib, socket, contextlib, json, urllib.request, urllib.error
import pandas as pd
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from io import BytesIO

from flask import (Flask, render_template, request, redirect, url_for,
                    session, flash, send_file, abort, g)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("RENDER") is not None,
)
DB = os.environ.get("DB_PATH") or os.path.join(os.path.dirname(__file__), "prestamos.db")
TASA_INTERES = 0.20  # 20% sobre el monto. Ej: $1.000.000 -> $1.200.000 a pagar
DEFAULT_ADMIN_PASSWORD = "admin123"


# --------------------------------------------------------------------------- DB
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS solicitudes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            cedula TEXT NOT NULL,
            telefono TEXT NOT NULL,
            correo TEXT,
            direccion TEXT NOT NULL,
            ingreso_mensual REAL NOT NULL,
            monto_solicitado REAL NOT NULL,
            num_cuotas INTEGER NOT NULL,
            estado TEXT NOT NULL DEFAULT 'Pendiente',
            viable INTEGER NOT NULL,
            cuota_estimada REAL NOT NULL,
            total_a_pagar REAL NOT NULL DEFAULT 0,
            monto_aprobado REAL,
            fecha_visita TEXT,
            notas TEXT,
            creado_en TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pagos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            solicitud_id INTEGER NOT NULL,
            monto REAL NOT NULL,
            fecha TEXT NOT NULL,
            notas TEXT,
            creado_en TEXT NOT NULL,
            FOREIGN KEY (solicitud_id) REFERENCES solicitudes(id) ON DELETE CASCADE
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS gastos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            concepto TEXT NOT NULL,
            monto REAL NOT NULL,
            fecha TEXT NOT NULL,
            creado_en TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            clave TEXT PRIMARY KEY,
            valor TEXT
        )
    """)
    # contraseña admin por defecto (hasheada)
    row = conn.execute("SELECT valor FROM settings WHERE clave='admin_password_hash'").fetchone()
    if not row:
        conn.execute("INSERT INTO settings (clave, valor) VALUES (?, ?)",
                      ("admin_password_hash", generate_password_hash(DEFAULT_ADMIN_PASSWORD)))
    cols = [c["name"] for c in conn.execute("PRAGMA table_info(solicitudes)").fetchall()]
    if "monto_aprobado" not in cols:
        conn.execute("ALTER TABLE solicitudes ADD COLUMN monto_aprobado REAL")
    if "puntaje" not in cols:
        conn.execute("ALTER TABLE solicitudes ADD COLUMN puntaje INTEGER NOT NULL DEFAULT 400")
    if "clave_hash" not in cols:
        conn.execute("ALTER TABLE solicitudes ADD COLUMN clave_hash TEXT")
    if "fecha_aprobacion" not in cols:
        conn.execute("ALTER TABLE solicitudes ADD COLUMN fecha_aprobacion TEXT")
    if "bloqueado" not in cols:
        conn.execute("ALTER TABLE solicitudes ADD COLUMN bloqueado INTEGER NOT NULL DEFAULT 0")
    if "mora_revisada" not in cols:
        conn.execute("ALTER TABLE solicitudes ADD COLUMN mora_revisada INTEGER NOT NULL DEFAULT 0")

    # genera clave de acceso (cédula) para clientes que ya existían sin clave_hash
    sin_clave = conn.execute("SELECT id, cedula FROM solicitudes WHERE clave_hash IS NULL").fetchall()
    for row in sin_clave:
        conn.execute("UPDATE solicitudes SET clave_hash=? WHERE id=?",
                      (generate_password_hash(row["cedula"]), row["id"]))

    for clave in ("smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_from", "brevo_api_key"):
        row = conn.execute("SELECT valor FROM settings WHERE clave=?", (clave,)).fetchone()
        if not row:
            conn.execute("INSERT INTO settings (clave, valor) VALUES (?, ?)", (clave, ""))
    conn.commit()
    conn.close()


def get_setting(clave, default=""):
    conn = get_db()
    row = conn.execute("SELECT valor FROM settings WHERE clave=?", (clave,)).fetchone()
    conn.close()
    return row["valor"] if row and row["valor"] is not None else default


def set_setting(clave, valor):
    conn = get_db()
    conn.execute("INSERT INTO settings (clave, valor) VALUES (?, ?) "
                  "ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor", (clave, valor))
    conn.commit()
    conn.close()


# --------------------------------------------------------------------------- Lógica de préstamo
def evaluar_viabilidad(ingreso_mensual, monto_solicitado, num_cuotas):
    """Tasa fija del 20% sobre el monto, pagada en cuotas semanales.
    Ej: $1.000.000 -> total $1.200.000 / 7 cuotas = $171.429 semanales.
    Se considera viable si la cuota semanal no supera el 35% del ingreso
    semanal (ingreso mensual / 4) y el monto no supera 3 veces el ingreso mensual."""
    total_a_pagar = monto_solicitado * (1 + TASA_INTERES)
    cuota = total_a_pagar / num_cuotas
    ingreso_semanal = ingreso_mensual / 4
    capacidad_pago = ingreso_semanal * 0.35
    tope_monto = ingreso_mensual * 3
    viable = cuota <= capacidad_pago and monto_solicitado <= tope_monto
    return viable, round(cuota, 2), round(total_a_pagar, 2)


def saldo_pendiente(conn, solicitud):
    abonado = conn.execute(
        "SELECT COALESCE(SUM(monto),0) AS s FROM pagos WHERE solicitud_id=?", (solicitud["id"],)
    ).fetchone()["s"]
    return round(solicitud["total_a_pagar"] - abonado, 2), round(abonado, 2)


PUNTAJE_MIN, PUNTAJE_MAX, PUNTAJE_INICIAL = 300, 850, 400


def estado_cuenta(solicitud, pagos):
    """Genera el plan de pagos semanal y marca en rojo ('atrasada') las cuotas
    cuyo vencimiento ya pasó hace más de 1 semana y aún no están cubiertas
    por el total abonado hasta la fecha."""
    cuotas = []
    if solicitud["estado"] != "Aprobado" or not solicitud["fecha_aprobacion"]:
        return cuotas

    fecha_base = datetime.strptime(solicitud["fecha_aprobacion"], "%Y-%m-%d")
    abonado_total = sum(p["monto"] for p in pagos)
    hoy = datetime.now()
    total_a_pagar = solicitud["total_a_pagar"]

    acumulado_esperado = 0
    for n in range(1, solicitud["num_cuotas"] + 1):
        vencimiento = fecha_base + timedelta(weeks=n)
        acumulado_esperado = min(acumulado_esperado + solicitud["cuota_estimada"], total_a_pagar)
        if abonado_total >= acumulado_esperado - 0.01:
            estado = "Pagada"
        elif hoy > vencimiento + timedelta(days=7):
            estado = "Atrasada"
        else:
            estado = "Pendiente"
        cuotas.append({
            "numero": n,
            "vencimiento": vencimiento.strftime("%Y-%m-%d"),
            "monto": round(solicitud["cuota_estimada"], 2),
            "estado": estado,
            "atrasada": estado == "Atrasada",
        })
    return cuotas


def recalcular_puntaje(conn, sid):
    """Ajusta el puntaje crediticio (estilo datacrédito, 300-850, inicia en 400)
    según la puntualidad de los abonos frente a las cuotas semanales."""
    s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (sid,)).fetchone()
    if s["estado"] != "Aprobado" or not s["fecha_aprobacion"]:
        return
    pagos = conn.execute(
        "SELECT * FROM pagos WHERE solicitud_id=? ORDER BY fecha ASC, id ASC", (sid,)
    ).fetchall()

    puntaje = PUNTAJE_INICIAL
    fecha_base = datetime.strptime(s["fecha_aprobacion"], "%Y-%m-%d")
    total_a_pagar = s["total_a_pagar"]
    hoy = datetime.now()
    acumulado_esperado = 0
    for n in range(1, s["num_cuotas"] + 1):
        vencimiento = fecha_base + timedelta(weeks=n)
        acumulado_esperado = min(acumulado_esperado + s["cuota_estimada"], total_a_pagar)
        acumulado_pagado = sum(
            p["monto"] for p in pagos if datetime.strptime(p["fecha"], "%Y-%m-%d") <= vencimiento
        )
        if acumulado_pagado >= acumulado_esperado - 0.01:
            puntaje += 10  # pago puntual (o anticipado)
        elif hoy > vencimiento:
            puntaje -= 15  # cuota atrasada
        # si la cuota aún no vence y no se ha cubierto, no afecta el puntaje todavía

    puntaje = max(PUNTAJE_MIN, min(PUNTAJE_MAX, puntaje))
    conn.execute("UPDATE solicitudes SET puntaje=? WHERE id=?", (puntaje, sid))
    conn.commit()


def verificar_fin_credito(conn, s):
    """Si el crédito ya llegó a su última cuota y quedó saldo pendiente, el
    cliente queda en mora: puntaje en 0 y bloqueado para nuevas solicitudes
    hasta que el administrador lo revise y le dé una nueva oportunidad."""
    if s["estado"] != "Aprobado" or not s["fecha_aprobacion"] or s["mora_revisada"]:
        return False
    fecha_base = datetime.strptime(s["fecha_aprobacion"], "%Y-%m-%d")
    fin_credito = fecha_base + timedelta(weeks=s["num_cuotas"])
    saldo, _ = saldo_pendiente(conn, s)
    if datetime.now() > fin_credito and saldo > 0.01:
        conn.execute("UPDATE solicitudes SET puntaje=0, bloqueado=1, mora_revisada=1 WHERE id=?", (s["id"],))
        conn.commit()
        return True
    return False


def credito_activo_no_renovable(conn, cedula):
    """Si el cliente tiene un crédito Aprobado con saldo pendiente, solo se
    permite una nueva solicitud cuando ya solo falta la última cuota por pagar.
    En cualquier otro caso (crédito recién iniciado o a mitad de pago) se
    informa que ya tiene un crédito activo."""
    activos = conn.execute(
        "SELECT * FROM solicitudes WHERE cedula=? AND estado='Aprobado' AND bloqueado=0", (cedula,)
    ).fetchall()
    for s in activos:
        saldo, _ = saldo_pendiente(conn, s)
        if saldo <= 0.01:
            continue  # ya está pagado, no cuenta como activo
        pagos = conn.execute("SELECT * FROM pagos WHERE solicitud_id=?", (s["id"],)).fetchall()
        cuotas = estado_cuenta(s, pagos)
        if not cuotas:
            return True
        pendientes = [c for c in cuotas if c["estado"] != "Pagada"]
        if len(pendientes) > 1 or (pendientes and pendientes[0]["numero"] != s["num_cuotas"]):
            return True  # tiene crédito activo, no solo la última cuota pendiente
    return False


def analisis_aumento(solicitud):
    """Sugiere si el cliente es elegible para un aumento de crédito según su puntaje."""
    if solicitud["estado"] != "Aprobado":
        return None
    puntaje = solicitud["puntaje"]
    monto_actual = solicitud["monto_aprobado"] or solicitud["monto_solicitado"]
    if puntaje >= 600:
        return {"elegible": True, "monto_sugerido": round(monto_actual * 1.5, -3),
                "mensaje": "Excelente historial de pago. Cliente elegible para aumento de crédito de hasta 50%."}
    if puntaje >= 500:
        return {"elegible": True, "monto_sugerido": round(monto_actual * 1.25, -3),
                "mensaje": "Buen historial de pago. Cliente elegible para aumento de crédito de hasta 25%."}
    return {"elegible": False, "monto_sugerido": monto_actual,
            "mensaje": "Aún no cumple los requisitos para un aumento de crédito."}


# --------------------------------------------------------------------------- Correo
@contextlib.contextmanager
def _forzar_ipv4():
    """Algunos hosts (ej. plan gratuito de Render) no tienen salida IPv6 y
    fallan con 'Network is unreachable' al resolver smtp.gmail.com vía AAAA.
    Forzamos getaddrinfo a devolver solo direcciones IPv4 durante el envío."""
    original = socket.getaddrinfo

    def solo_ipv4(host, port, family=0, type=0, proto=0, flags=0):
        return original(host, port, socket.AF_INET, type, proto, flags)

    socket.getaddrinfo = solo_ipv4
    try:
        yield
    finally:
        socket.getaddrinfo = original


def _enviar_correo_brevo(destinatario, asunto, cuerpo):
    """Envía el correo vía la API HTTP de Brevo (puerto 443), que funciona en
    hosts gratuitos como Render donde los puertos SMTP salientes están bloqueados."""
    api_key = get_setting("brevo_api_key")
    remitente = get_setting("smtp_from") or get_setting("smtp_user")
    if not (api_key and remitente):
        return None  # no configurado, intenta el siguiente método
    data = json.dumps({
        "sender": {"email": remitente, "name": "Créditos Crecer"},
        "to": [{"email": destinatario}],
        "subject": asunto,
        "textContent": cuerpo,
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.brevo.com/v3/smtp/email", data=data, method="POST",
        headers={"api-key": api_key, "Content-Type": "application/json", "Accept": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            if 200 <= resp.status < 300:
                return True, "Correo enviado correctamente."
            return False, f"Brevo respondió con código {resp.status}."
    except urllib.error.HTTPError as exc:
        detalle = exc.read().decode("utf-8", "ignore")
        return False, f"No se pudo enviar el correo (Brevo): HTTP {exc.code} - {detalle}"
    except Exception as exc:
        return False, f"No se pudo enviar el correo (Brevo): {exc}"


def enviar_correo(destinatario, asunto, cuerpo):
    if not destinatario:
        return False, "El cliente no registró un correo electrónico."

    resultado_brevo = _enviar_correo_brevo(destinatario, asunto, cuerpo)
    if resultado_brevo is not None:
        return resultado_brevo

    host = get_setting("smtp_host")
    port = get_setting("smtp_port")
    user = get_setting("smtp_user")
    password = get_setting("smtp_password")
    remitente = get_setting("smtp_from") or user
    if not (host and port and user and password):
        return False, "La configuración de correo no está completa (configura Brevo o SMTP)."
    try:
        msg = MIMEText(cuerpo, "plain", "utf-8")
        msg["Subject"] = asunto
        msg["From"] = remitente
        msg["To"] = destinatario
        port_int = int(port)
        with _forzar_ipv4():
            if port_int == 465:
                server = smtplib.SMTP_SSL(host, port_int, timeout=15)
            else:
                server = smtplib.SMTP(host, port_int, timeout=15)
            with server:
                if port_int != 465:
                    server.starttls()
                server.login(user, password)
                server.sendmail(remitente, [destinatario], msg.as_string())
        return True, "Correo enviado correctamente."
    except Exception as exc:
        return False, f"No se pudo enviar el correo: {exc}"


def fmt_money(n):
    return "${:,.0f}".format(n or 0)


app.jinja_env.filters["money"] = fmt_money


# --------------------------------------------------------------------------- Seguridad (CSRF + sesión admin)
@app.before_request
def seguridad():
    # Genera un token CSRF por sesión
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(16)

    if request.method == "POST":
        token = session.get("csrf_token")
        enviado = request.form.get("csrf_token")
        if not token or not enviado or not secrets.compare_digest(token, enviado):
            abort(400, "Token de seguridad inválido. Recarga la página e inténtalo de nuevo.")

    # Protege todo /admin (excepto login) tras sesión iniciada
    if request.path.startswith("/admin") and request.path != "/admin/login":
        if not session.get("admin"):
            return redirect(url_for("admin_login"))

    # Protege el portal de clientes
    if request.path.startswith("/cliente") and request.path != "/cliente/login":
        if not session.get("cliente_id"):
            return redirect(url_for("cliente_login"))


@app.context_processor
def inject_csrf():
    return {"csrf_token": session.get("csrf_token", "")}


@app.after_request
def cabeceras_seguridad(resp):
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["X-Frame-Options"] = "DENY"
    resp.headers["Referrer-Policy"] = "no-referrer"
    return resp


# --------------------------------------------------------------------------- Rutas públicas
@app.route("/")
def registro():
    return render_template("registro.html")


@app.route("/registro", methods=["POST"])
def registro_post():
    nombre = request.form["nombre"].strip()
    cedula = request.form["cedula"].strip()
    telefono = request.form["telefono"].strip()
    correo = request.form.get("correo", "").strip()
    direccion = request.form["direccion"].strip()
    ingreso_mensual = float(request.form["ingreso_mensual"])
    monto_solicitado = float(request.form["monto_solicitado"])
    num_cuotas = int(request.form["num_cuotas"])

    viable, cuota, total_a_pagar = evaluar_viabilidad(ingreso_mensual, monto_solicitado, num_cuotas)

    conn = get_db()

    bloqueo = conn.execute(
        "SELECT id FROM solicitudes WHERE cedula=? AND bloqueado=1 ORDER BY id DESC LIMIT 1", (cedula,)
    ).fetchone()
    credito_activo = False if bloqueo else credito_activo_no_renovable(conn, cedula)
    estado_inicial = "Rechazado" if (bloqueo or credito_activo) else "Pendiente"
    viable_final = 0 if (bloqueo or credito_activo) else int(viable)

    conn.execute(
        """INSERT INTO solicitudes
           (nombre, cedula, telefono, correo, direccion, ingreso_mensual, monto_solicitado,
            num_cuotas, estado, viable, cuota_estimada, total_a_pagar, puntaje, clave_hash, creado_en)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (nombre, cedula, telefono, correo, direccion, ingreso_mensual, monto_solicitado,
         num_cuotas, estado_inicial, viable_final, cuota, total_a_pagar, PUNTAJE_INICIAL,
         generate_password_hash(cedula), datetime.now().strftime("%Y-%m-%d %H:%M")),
    )
    conn.commit()
    conn.close()

    if bloqueo:
        return render_template("gracias.html", nombre=nombre, bloqueado=True)

    if credito_activo:
        return render_template("gracias.html", nombre=nombre, credito_activo=True)

    return render_template("gracias.html", nombre=nombre, viable=viable, cuota=cuota,
                            total_a_pagar=total_a_pagar, num_cuotas=num_cuotas)


# --------------------------------------------------------------------------- Login / sesión admin
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        password_hash = get_setting("admin_password_hash")
        if password_hash and check_password_hash(password_hash, request.form.get("password", "")):
            session.clear()
            session["admin"] = True
            session["csrf_token"] = secrets.token_hex(16)
            return redirect(url_for("admin_dashboard"))
        flash("Contraseña incorrecta")
    return render_template("admin_login.html")


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


# --------------------------------------------------------------------------- Dashboard
@app.route("/admin")
def admin_dashboard():
    conn = get_db()
    solicitudes = conn.execute("SELECT * FROM solicitudes ORDER BY creado_en DESC").fetchall()

    detalle = []
    total_prestado = total_por_cobrar = total_abonado = 0
    for s in solicitudes:
        if verificar_fin_credito(conn, s):
            s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (s["id"],)).fetchone()
        saldo, abonado = saldo_pendiente(conn, s)
        d = dict(s)
        d["saldo"] = saldo
        d["abonado"] = abonado
        pagos_s = conn.execute("SELECT * FROM pagos WHERE solicitud_id=?", (s["id"],)).fetchall()
        d["cuotas"] = estado_cuenta(s, pagos_s)
        d["atrasado"] = any(c["atrasada"] for c in d["cuotas"])
        d["aumento"] = analisis_aumento(s)
        detalle.append(d)
        if s["estado"] == "Aprobado":
            total_prestado += s["monto_aprobado"] if s["monto_aprobado"] is not None else s["monto_solicitado"]
            total_por_cobrar += saldo
        total_abonado += abonado

    total_gastos = conn.execute("SELECT COALESCE(SUM(monto),0) AS s FROM gastos").fetchone()["s"]
    conn.close()

    total = len(solicitudes)
    viables = sum(1 for s in solicitudes if s["viable"])
    aprobados = sum(1 for s in solicitudes if s["estado"] == "Aprobado")
    pendientes = sum(1 for s in solicitudes if s["estado"] == "Pendiente")

    return render_template(
        "admin.html",
        solicitudes=detalle,
        total=total, viables=viables, aprobados=aprobados, pendientes=pendientes,
        total_prestado=total_prestado, total_por_cobrar=total_por_cobrar,
        total_abonado=total_abonado, total_gastos=total_gastos,
    )


# --------------------------------------------------------------------------- Estado (Pendiente / Rechazado)
@app.route("/admin/solicitud/<int:sid>/estado", methods=["POST"])
def admin_estado(sid):
    estado = request.form["estado"]
    if estado == "Aprobado":
        # La aprobación se hace por /aprobar para poder ajustar el monto
        return redirect(url_for("admin_dashboard"))
    conn = get_db()
    conn.execute("UPDATE solicitudes SET estado=? WHERE id=?", (estado, sid))
    conn.commit()
    conn.close()
    flash("Estado actualizado.")
    return redirect(url_for("admin_dashboard"))


# --------------------------------------------------------------------------- Aprobación (monto ajustable + correo)
@app.route("/admin/solicitud/<int:sid>/aprobar", methods=["POST"])
def admin_aprobar(sid):
    conn = get_db()
    s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (sid,)).fetchone()
    if not s:
        conn.close()
        abort(404)

    monto_aprobado = float(request.form.get("monto_aprobado") or s["monto_solicitado"])
    num_cuotas = int(request.form.get("num_cuotas_aprobadas") or s["num_cuotas"])
    if monto_aprobado <= 0:
        monto_aprobado = s["monto_solicitado"]
    if num_cuotas <= 0:
        num_cuotas = s["num_cuotas"]

    total_a_pagar = round(monto_aprobado * (1 + TASA_INTERES), 2)
    cuota = round(total_a_pagar / num_cuotas, 2)

    conn.execute(
        """UPDATE solicitudes
           SET estado='Aprobado', monto_aprobado=?, num_cuotas=?, total_a_pagar=?, cuota_estimada=?,
               fecha_aprobacion=?, bloqueado=0, mora_revisada=0
           WHERE id=?""",
        (monto_aprobado, num_cuotas, total_a_pagar, cuota, datetime.now().strftime("%Y-%m-%d"), sid),
    )
    conn.commit()
    conn.close()

    cuerpo = (
        f"Hola {s['nombre']},\n\n"
        f"¡Buenas noticias! Tu solicitud de préstamo ha sido APROBADA"
        + (f" por un monto de {fmt_money(monto_aprobado)}"
           if monto_aprobado != s["monto_solicitado"]
           else f" por {fmt_money(monto_aprobado)}")
        + ".\n\n"
        f"Detalles del préstamo:\n"
        f"  - Total a pagar: {fmt_money(total_a_pagar)}\n"
        f"  - Número de cuotas: {num_cuotas}\n"
        f"  - Valor de cada cuota: {fmt_money(cuota)}\n\n"
        f"Pronto te contactaremos para coordinar la entrega y la primera visita de cobro.\n\n"
        f"Puedes consultar tu estado de cuenta en cualquier momento ingresando a nuestro portal "
        f"de clientes con tu número de cédula como usuario y contraseña inicial "
        f"(te recomendamos cambiarla luego).\n\n"
        f"Créditos Crecer\n"
        f"Confianza · Seguridad · Rápido y fácil"
    )
    ok, msg = enviar_correo(s["correo"], "Tu préstamo fue aprobado - Créditos Crecer", cuerpo)
    if ok:
        flash("Préstamo aprobado y correo enviado al cliente.")
    else:
        flash(f"Préstamo aprobado. Aviso: {msg}")

    return redirect(url_for("admin_dashboard"))


# --------------------------------------------------------------------------- [TEMPORAL] Simular fin de crédito vencido (para pruebas)
@app.route("/admin/solicitud/<int:sid>/simular_mora", methods=["POST"])
def admin_simular_mora(sid):
    conn = get_db()
    s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (sid,)).fetchone()
    if not s or s["estado"] != "Aprobado":
        conn.close()
        flash("El préstamo debe estar Aprobado para simular esto.")
        return redirect(url_for("admin_dashboard"))
    fecha_pasada = (datetime.now() - timedelta(weeks=s["num_cuotas"] + 1)).strftime("%Y-%m-%d")
    conn.execute("UPDATE solicitudes SET fecha_aprobacion=?, bloqueado=0, mora_revisada=0 WHERE id=?",
                  (fecha_pasada, sid))
    conn.commit()
    conn.close()
    flash("Simulación aplicada: este crédito ahora aparece vencido sin pagar.")
    return redirect(url_for("admin_dashboard"))


# --------------------------------------------------------------------------- Desbloquear cliente (nueva oportunidad)
@app.route("/admin/solicitud/<int:sid>/desbloquear", methods=["POST"])
def admin_desbloquear(sid):
    nuevo_puntaje = int(request.form.get("nuevo_puntaje") or PUNTAJE_INICIAL)
    nuevo_puntaje = max(PUNTAJE_MIN, min(PUNTAJE_MAX, nuevo_puntaje))
    conn = get_db()
    conn.execute("UPDATE solicitudes SET bloqueado=0, puntaje=? WHERE id=?", (nuevo_puntaje, sid))
    conn.commit()
    conn.close()
    flash("Cliente desbloqueado con nuevo puntaje.")
    return redirect(url_for("admin_dashboard"))


# --------------------------------------------------------------------------- Visitas
@app.route("/admin/solicitud/<int:sid>/visita", methods=["POST"])
def admin_visita(sid):
    fecha_visita = request.form["fecha_visita"]
    notas = request.form.get("notas", "")
    conn = get_db()
    conn.execute("UPDATE solicitudes SET fecha_visita=?, notas=? WHERE id=?", (fecha_visita, notas, sid))
    conn.commit()
    conn.close()
    flash("Visita programada.")
    return redirect(url_for("admin_dashboard"))


# --------------------------------------------------------------------------- Pagos / abonos (envía correo)
@app.route("/admin/solicitud/<int:sid>/pago", methods=["POST"])
def admin_pago(sid):
    monto = float(request.form["monto"])
    fecha = request.form.get("fecha") or datetime.now().strftime("%Y-%m-%d")
    notas = request.form.get("notas", "")

    conn = get_db()
    s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (sid,)).fetchone()
    if not s:
        conn.close()
        abort(404)

    conn.execute(
        "INSERT INTO pagos (solicitud_id, monto, fecha, notas, creado_en) VALUES (?, ?, ?, ?, ?)",
        (sid, monto, fecha, notas, datetime.now().strftime("%Y-%m-%d %H:%M")),
    )
    conn.commit()
    recalcular_puntaje(conn, sid)
    saldo, abonado = saldo_pendiente(conn, s)
    conn.close()

    cuerpo = (
        f"Hola {s['nombre']},\n\n"
        f"Hemos registrado un abono a tu préstamo.\n\n"
        f"Detalle del abono:\n"
        f"  - Valor abonado: {fmt_money(monto)}\n"
        f"  - Fecha: {fecha}\n\n"
        f"Resumen de tu préstamo:\n"
        f"  - Cuota semanal aproximada: {fmt_money(s['cuota_estimada'])}\n"
        f"  - Total abonado hasta hoy: {fmt_money(abonado)}\n"
        f"  - Saldo pendiente: {fmt_money(saldo)}\n\n"
        f"Gracias por tu pago puntual.\n\n"
        f"Créditos Crecer\n"
        f"Confianza · Seguridad · Rápido y fácil"
    )
    ok, msg = enviar_correo(s["correo"], "Confirmación de abono - Créditos Crecer", cuerpo)
    if ok:
        flash("Abono registrado y correo de confirmación enviado.")
    else:
        flash(f"Abono registrado. Aviso: {msg}")

    return redirect(url_for("admin_dashboard"))


# --------------------------------------------------------------------------- Historial de abonos (ver / corregir / eliminar)
@app.route("/admin/solicitud/<int:sid>/abonos")
def admin_abonos(sid):
    conn = get_db()
    s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (sid,)).fetchone()
    if not s:
        conn.close()
        abort(404)
    pagos = conn.execute(
        "SELECT * FROM pagos WHERE solicitud_id=? ORDER BY fecha DESC, id DESC", (sid,)
    ).fetchall()
    saldo, abonado = saldo_pendiente(conn, s)
    conn.close()
    return render_template("abonos.html", s=s, pagos=pagos, saldo=saldo, abonado=abonado)


@app.route("/admin/pago/<int:pid>/editar", methods=["POST"])
def admin_pago_editar(pid):
    monto = float(request.form["monto"])
    fecha = request.form["fecha"]
    notas = request.form.get("notas", "")
    conn = get_db()
    pago = conn.execute("SELECT solicitud_id FROM pagos WHERE id=?", (pid,)).fetchone()
    if not pago:
        conn.close()
        abort(404)
    conn.execute("UPDATE pagos SET monto=?, fecha=?, notas=? WHERE id=?", (monto, fecha, notas, pid))
    conn.commit()
    sid = pago["solicitud_id"]
    recalcular_puntaje(conn, sid)
    conn.close()
    flash("Abono corregido.")
    return redirect(url_for("admin_abonos", sid=sid))


@app.route("/admin/pago/<int:pid>/eliminar", methods=["POST"])
def admin_pago_eliminar(pid):
    conn = get_db()
    pago = conn.execute("SELECT solicitud_id FROM pagos WHERE id=?", (pid,)).fetchone()
    if not pago:
        conn.close()
        abort(404)
    sid = pago["solicitud_id"]
    conn.execute("DELETE FROM pagos WHERE id=?", (pid,))
    conn.commit()
    recalcular_puntaje(conn, sid)
    conn.close()
    flash("Abono eliminado.")
    return redirect(url_for("admin_abonos", sid=sid))


# --------------------------------------------------------------------------- Portal de clientes
@app.route("/cliente/login", methods=["GET", "POST"])
def cliente_login():
    if request.method == "POST":
        cedula = request.form.get("cedula", "").strip()
        clave = request.form.get("clave", "")
        conn = get_db()
        s = conn.execute("SELECT * FROM solicitudes WHERE cedula=?", (cedula,)).fetchone()
        conn.close()
        if s and s["clave_hash"] and check_password_hash(s["clave_hash"], clave):
            session.clear()
            session["cliente_id"] = s["id"]
            session["csrf_token"] = secrets.token_hex(16)
            return redirect(url_for("cliente_dashboard"))
        flash("Cédula o contraseña incorrecta.")
    return render_template("cliente_login.html")


@app.route("/cliente/logout", methods=["POST"])
def cliente_logout():
    session.clear()
    return redirect(url_for("cliente_login"))


@app.route("/cliente")
def cliente_dashboard():
    conn = get_db()
    s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (session["cliente_id"],)).fetchone()
    if not s:
        session.clear()
        conn.close()
        return redirect(url_for("cliente_login"))
    if verificar_fin_credito(conn, s):
        s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (s["id"],)).fetchone()
    pagos = conn.execute(
        "SELECT * FROM pagos WHERE solicitud_id=? ORDER BY fecha DESC, id DESC", (s["id"],)
    ).fetchall()
    saldo, abonado = saldo_pendiente(conn, s)
    conn.close()
    cuotas = estado_cuenta(s, pagos)
    aumento = analisis_aumento(s)
    return render_template("cliente.html", s=s, pagos=pagos, saldo=saldo, abonado=abonado,
                            cuotas=cuotas, aumento=aumento)


@app.route("/cliente/clave", methods=["GET", "POST"])
def cliente_clave():
    if request.method == "POST":
        actual = request.form.get("actual", "")
        nueva = request.form.get("nueva", "")
        confirmar = request.form.get("confirmar", "")
        conn = get_db()
        s = conn.execute("SELECT * FROM solicitudes WHERE id=?", (session["cliente_id"],)).fetchone()
        if not check_password_hash(s["clave_hash"], actual):
            flash("La contraseña actual es incorrecta.")
        elif len(nueva) < 4:
            flash("La nueva contraseña debe tener al menos 4 caracteres.")
        elif nueva != confirmar:
            flash("La confirmación no coincide con la nueva contraseña.")
        else:
            conn.execute("UPDATE solicitudes SET clave_hash=? WHERE id=?",
                          (generate_password_hash(nueva), s["id"]))
            conn.commit()
            flash("Contraseña actualizada correctamente.")
            conn.close()
            return redirect(url_for("cliente_dashboard"))
        conn.close()
    return render_template("cliente_clave.html")


# --------------------------------------------------------------------------- Eliminar solicitud
@app.route("/admin/solicitud/<int:sid>/eliminar", methods=["POST"])
def admin_eliminar(sid):
    conn = get_db()
    conn.execute("DELETE FROM pagos WHERE solicitud_id=?", (sid,))
    conn.execute("DELETE FROM solicitudes WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    flash("Solicitud eliminada.")
    return redirect(url_for("admin_dashboard"))


# --------------------------------------------------------------------------- Gastos administrativos
@app.route("/admin/gastos")
def admin_gastos():
    conn = get_db()
    gastos = conn.execute("SELECT * FROM gastos ORDER BY fecha DESC, id DESC").fetchall()
    total_gastos = conn.execute("SELECT COALESCE(SUM(monto),0) AS s FROM gastos").fetchone()["s"]
    conn.close()
    return render_template("gastos.html", gastos=gastos, total_gastos=total_gastos)


@app.route("/admin/gastos/nuevo", methods=["POST"])
def admin_gastos_nuevo():
    concepto = request.form["concepto"].strip()
    monto = float(request.form["monto"])
    fecha = request.form.get("fecha") or datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    conn.execute("INSERT INTO gastos (concepto, monto, fecha, creado_en) VALUES (?, ?, ?, ?)",
                  (concepto, monto, fecha, datetime.now().strftime("%Y-%m-%d %H:%M")))
    conn.commit()
    conn.close()
    flash("Gasto registrado.")
    return redirect(url_for("admin_gastos"))


@app.route("/admin/gastos/<int:gid>/eliminar", methods=["POST"])
def admin_gastos_eliminar(gid):
    conn = get_db()
    conn.execute("DELETE FROM gastos WHERE id=?", (gid,))
    conn.commit()
    conn.close()
    flash("Gasto eliminado.")
    return redirect(url_for("admin_gastos"))


# --------------------------------------------------------------------------- Cambiar clave
@app.route("/admin/clave", methods=["GET", "POST"])
def admin_clave():
    if request.method == "POST":
        actual = request.form.get("actual", "")
        nueva = request.form.get("nueva", "")
        confirmar = request.form.get("confirmar", "")
        password_hash = get_setting("admin_password_hash")
        if not check_password_hash(password_hash, actual):
            flash("La contraseña actual es incorrecta.")
        elif len(nueva) < 6:
            flash("La nueva contraseña debe tener al menos 6 caracteres.")
        elif nueva != confirmar:
            flash("La confirmación no coincide con la nueva contraseña.")
        else:
            set_setting("admin_password_hash", generate_password_hash(nueva))
            flash("Contraseña actualizada correctamente.")
            return redirect(url_for("admin_dashboard"))
    return render_template("clave.html")


# --------------------------------------------------------------------------- Configuración SMTP
@app.route("/admin/configuracion", methods=["GET", "POST"])
def admin_configuracion():
    if request.method == "POST":
        if "smtp_host" in request.form:
            set_setting("smtp_host", request.form.get("smtp_host", "").strip())
            set_setting("smtp_port", request.form.get("smtp_port", "").strip())
            set_setting("smtp_user", request.form.get("smtp_user", "").strip())
            nueva_pass = request.form.get("smtp_password", "")
            if nueva_pass:
                set_setting("smtp_password", nueva_pass)
        nueva_api_key = request.form.get("brevo_api_key", "").strip()
        if nueva_api_key:
            set_setting("brevo_api_key", nueva_api_key)
        if request.form.get("smtp_from", "").strip():
            set_setting("smtp_from", request.form.get("smtp_from", "").strip())
        flash("Configuración de correo guardada.")
        return redirect(url_for("admin_configuracion"))

    config = {
        "smtp_host": get_setting("smtp_host"),
        "smtp_port": get_setting("smtp_port"),
        "smtp_user": get_setting("smtp_user"),
        "smtp_from": get_setting("smtp_from"),
        "smtp_password_set": bool(get_setting("smtp_password")),
        "brevo_api_key_set": bool(get_setting("brevo_api_key")),
    }
    return render_template("configuracion.html", config=config)


@app.route("/admin/configuracion/probar", methods=["POST"])
def admin_configuracion_probar():
    destino = request.form.get("correo_prueba", "").strip()
    cuerpo = (
        "Este es un correo de prueba enviado desde el panel de Créditos Crecer.\n\n"
        "Si lo recibiste, la configuración SMTP funciona correctamente."
    )
    ok, msg = enviar_correo(destino, "Correo de prueba - Créditos Crecer", cuerpo)
    flash(("Éxito: " if ok else "Error: ") + msg)
    return redirect(url_for("admin_configuracion"))


# --------------------------------------------------------------------------- Importar desde Excel
def _to_float(v):
    try:
        if v is None:
            return 0.0
        import math
        if isinstance(v, float) and math.isnan(v):
            return 0.0
        return float(v)
    except Exception:
        return 0.0


def _leer_hoja_excel(contenido_bytes, nombre_hoja):
    """Lee una hoja buscando la fila que contiene los encabezados reales."""
    from io import BytesIO as _BIO
    df_raw = pd.read_excel(_BIO(contenido_bytes), sheet_name=nombre_hoja, header=None)
    for i, row in df_raw.iterrows():
        vals = [str(v).strip() for v in row.values]
        if any(m in vals for m in ('COD CLIENTE', 'NRO DOC')):
            # Normalizar nombres de columna (quitar espacios múltiples y nans)
            cols = [' '.join(str(v).split()) if str(v).strip() not in ('nan', '') else None
                    for v in row.values]
            data = df_raw.iloc[i + 1:].copy()
            data.columns = cols
            return data.dropna(how='all').reset_index(drop=True)
    return pd.DataFrame()


@app.route("/admin/importar", methods=["GET", "POST"])
def admin_importar():
    if request.method == "GET":
        return render_template("importar.html")

    archivo = request.files.get("archivo")
    if not archivo or not archivo.filename:
        flash("Selecciona un archivo Excel (.xlsx) para importar.")
        return redirect(url_for("admin_importar"))

    try:
        from io import BytesIO as _BIO
        contenido = archivo.read()

        xls = pd.ExcelFile(_BIO(contenido))
        sheets_upper = {s.strip().upper(): s for s in xls.sheet_names}

        def encontrar_hojas(*keywords):
            """Devuelve TODAS las hojas cuyos nombres contienen alguna de las keywords."""
            resultado = []
            for kw in keywords:
                for upper, real in sheets_upper.items():
                    if kw in upper and real not in resultado:
                        resultado.append(real)
            return resultado

        def combinar_hojas(hojas):
            frames = []
            for h in hojas:
                df = _leer_hoja_excel(contenido, h)
                if not df.empty:
                    frames.append(df)
            return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

        hojas_clientes = encontrar_hojas('CLIENTE')
        hojas_docs = encontrar_hojas('DOCUMENTO', 'PRESTAMO', 'CREDITO')
        hojas_abonos = encontrar_hojas('ABONO', 'PAGO')

        if not hojas_docs:
            flash("No se encontró la hoja de documentos/préstamos. Verifica que el archivo tenga una hoja llamada DOCUMENTOS.")
            return redirect(url_for("admin_importar"))

        df_clientes = combinar_hojas(hojas_clientes)
        df_docs = combinar_hojas(hojas_docs)
        df_abonos = combinar_hojas(hojas_abonos)

        # Diccionario COD CLIENTE → {telefono, correo}
        info_cliente = {}
        if not df_clientes.empty:
            for _, row in df_clientes.iterrows():
                cod = str(row.get('COD CLIENTE') or '').strip()
                if not cod or cod == 'nan':
                    continue
                tel = str(row.get('CELULAR') or '').strip()
                cor = str(row.get('CORREO') or '').strip()
                # Limpiar valores "No refiere" / NaN
                tel = '' if tel.lower() in ('no refiere', 'nan', '') else tel.replace('.0', '')
                cor = '' if cor.lower() in ('no refiere', 'nan', '') else cor
                info_cliente[cod] = {'telefono': tel, 'correo': cor}

        conn = get_db()
        importados = omitidos = pagos_importados = 0
        nro_a_sid = {}  # NRO DOC → solicitud_id

        for _, row in df_docs.iterrows():
            nro_doc = str(row.get('NRO DOC') or '').strip()
            if not nro_doc or nro_doc == 'nan':
                continue
            nombre = str(row.get('CLIENTE') or '').strip()
            if not nombre or nombre == 'nan':
                continue

            # ¿Ya existe en la BD?
            existente = conn.execute(
                "SELECT id FROM solicitudes WHERE cedula=? AND estado='Aprobado'", (nro_doc,)
            ).fetchone()
            if existente:
                nro_a_sid[nro_doc] = existente['id']
                omitidos += 1
                continue

            cod_cliente = str(row.get('COD CLIENTE') or '').strip()
            info = info_cliente.get(cod_cliente, {})

            telefono = info.get('telefono') or 'Sin teléfono'
            correo = info.get('correo') or ''
            direccion = str(row.get('UBICACIÓN') or row.get('UBICACION') or '').strip()
            direccion = direccion if direccion not in ('nan', '') else 'Sin dirección'

            monto_real = _to_float(row.get('VALOR REAL PRESTADO'))
            total_a_pagar = _to_float(row.get('VALOR A COBRAR'))
            if monto_real <= 0:
                monto_real = round(total_a_pagar / 1.2, 2)
            dias = _to_float(row.get('DÍAS DE CRÉDITO') or row.get('DIAS DE CREDITO')) or 42
            num_cuotas = max(1, round(dias / 7))
            cuota_estimada = round(total_a_pagar / num_cuotas, 2) if total_a_pagar else 0

            fecha_raw = row.get('FECHA DE EMISIÓN') or row.get('FECHA DE EMISION')
            try:
                fecha_aprobacion = pd.Timestamp(fecha_raw).strftime('%Y-%m-%d')
            except Exception:
                fecha_aprobacion = datetime.now().strftime('%Y-%m-%d')

            ahora = datetime.now().strftime('%Y-%m-%d %H:%M')
            cur = conn.execute(
                """INSERT INTO solicitudes
                   (nombre, cedula, telefono, correo, direccion, ingreso_mensual,
                    monto_solicitado, num_cuotas, estado, viable, cuota_estimada,
                    total_a_pagar, monto_aprobado, puntaje, clave_hash,
                    fecha_aprobacion, bloqueado, mora_revisada, creado_en)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (nombre, nro_doc, telefono, correo, direccion,
                 0, monto_real, num_cuotas, 'Aprobado', 1,
                 cuota_estimada, total_a_pagar, monto_real, PUNTAJE_INICIAL,
                 generate_password_hash(nro_doc),
                 fecha_aprobacion, 0, 0, ahora)
            )
            conn.commit()
            nro_a_sid[nro_doc] = cur.lastrowid
            importados += 1

        # Importar abonos
        if not df_abonos.empty:
            for _, row in df_abonos.iterrows():
                nro_doc = str(row.get('NRO DOC') or '').strip()
                if not nro_doc or nro_doc == 'nan':
                    continue
                # Soportar NRO DOC compuesto como "1033680536-01"
                base = nro_doc.split('-')[0]
                sid = nro_a_sid.get(nro_doc) or nro_a_sid.get(base)
                if not sid:
                    continue
                valor = _to_float(row.get('VALOR ABONADO'))
                if valor <= 0:
                    continue
                fecha_raw = row.get('FECHA DE ABONO')
                try:
                    fecha_str = pd.Timestamp(fecha_raw).strftime('%Y-%m-%d')
                except Exception:
                    continue
                # Evitar duplicados: si ya existe un pago igual, omitir
                dup = conn.execute(
                    "SELECT id FROM pagos WHERE solicitud_id=? AND monto=? AND fecha=?",
                    (sid, valor, fecha_str)
                ).fetchone()
                if dup:
                    continue
                ahora = datetime.now().strftime('%Y-%m-%d %H:%M')
                conn.execute(
                    "INSERT INTO pagos (solicitud_id, monto, fecha, notas, creado_en) VALUES (?,?,?,?,?)",
                    (sid, valor, fecha_str, 'Importado desde Excel', ahora)
                )
                pagos_importados += 1
            conn.commit()

        # Recalcular puntaje de todos los importados
        for sid in set(nro_a_sid.values()):
            recalcular_puntaje(conn, sid)
        conn.close()

        flash(f"✓ Importación completada: {importados} créditos nuevos, "
              f"{pagos_importados} abonos importados, {omitidos} ya existían (omitidos).")
        return redirect(url_for("admin_dashboard"))

    except Exception as exc:
        flash(f"Error al procesar el archivo: {exc}")
        return redirect(url_for("admin_importar"))


# --------------------------------------------------------------------------- Exportar a Excel
@app.route("/admin/exportar")
def admin_exportar():
    conn = get_db()
    solicitudes = conn.execute("SELECT * FROM solicitudes ORDER BY creado_en DESC").fetchall()
    pagos = conn.execute("""SELECT p.*, s.nombre, s.cedula FROM pagos p
                             JOIN solicitudes s ON s.id = p.solicitud_id
                             ORDER BY p.fecha DESC""").fetchall()
    gastos = conn.execute("SELECT * FROM gastos ORDER BY fecha DESC").fetchall()
    conn.close()

    wb = Workbook()

    ws = wb.active
    ws.title = "Solicitudes"
    ws.append(["ID", "Nombre", "Cédula", "Teléfono", "Correo", "Dirección",
               "Ingreso mensual", "Monto solicitado", "N° cuotas", "Cuota semanal",
               "Total a pagar", "Abonado", "Saldo pendiente", "Viable", "Estado",
               "Fecha visita", "Notas", "Creado en"])
    conn = get_db()
    for s in solicitudes:
        saldo, abonado = saldo_pendiente(conn, s)
        ws.append([s["id"], s["nombre"], s["cedula"], s["telefono"], s["correo"], s["direccion"],
                   s["ingreso_mensual"], s["monto_solicitado"], s["num_cuotas"], s["cuota_estimada"],
                   s["total_a_pagar"], abonado, saldo, "Sí" if s["viable"] else "No", s["estado"],
                   s["fecha_visita"], s["notas"], s["creado_en"]])
    conn.close()

    ws2 = wb.create_sheet("Abonos")
    ws2.append(["ID", "Solicitud ID", "Cliente", "Cédula", "Monto", "Fecha", "Notas", "Registrado en"])
    for p in pagos:
        ws2.append([p["id"], p["solicitud_id"], p["nombre"], p["cedula"], p["monto"], p["fecha"],
                    p["notas"], p["creado_en"]])

    ws3 = wb.create_sheet("Gastos administrativos")
    ws3.append(["ID", "Concepto", "Monto", "Fecha", "Registrado en"])
    for g_ in gastos:
        ws3.append([g_["id"], g_["concepto"], g_["monto"], g_["fecha"], g_["creado_en"]])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            largo = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(largo + 2, 45)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_archivo = f"creditos_crecer_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nombre_archivo,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
